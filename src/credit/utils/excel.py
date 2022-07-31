import numpy as np
import time
import openpyxl
import os
import pandas as pd

from copy import copy
from datetime import datetime
from typing import List

# package imports
from . import paths

# --------------------------------------------------------------------------------
# Parametros
# --------------------------------------------------------------------------------
COLOR_HEADER = '173E89'
COLOR_BLANCO = 'FFFFFFFF'
COLOR_NEGRO = 'FF000000'


# --------------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------------
def save_dataframe(df, filename, sheetname, index=True, header=True,
                   title=None, currency_columns=None, integer_columns=None,
                   pct_columns=None, startcol=0, startrow=0,
                   color_index=False, skip_nan=False, sleep=None,
                   delete_workbook=False, delete_worksheet=False):
    if len(sheetname) > 31:
        sheetname = sheetname[:31]
        
    if title is not None:
        startrow += 1
        
    if "/" in filename:
        paths.mkdir("/".join(filename.split("/")[:-1]))
        
    df = df.reset_index(drop=not index).copy()
    
    for _ in range(10):
        try:
            if delete_workbook or not os.path.exists(filename):
                df.to_excel(filename, sheetname, index=False, header=header,
                            startrow=startrow, startcol=startcol)
                
            wb = openpyxl.load_workbook(filename)
            break
        except PermissionError as e:
            time.sleep(0.5)
            pass
    else:
        raise e

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        writer.book = wb
        writer.sheets = {ws.title:ws for ws in wb.worksheets}
        
        header_fill = openpyxl.styles.PatternFill(fgColor=COLOR_HEADER, fill_type="solid")
        header_font = openpyxl.styles.Font(name="Open Sans", size=11, bold=True, color=COLOR_BLANCO)
        
        border_side = openpyxl.styles.Side(border_style="thin", color=COLOR_NEGRO)
        border = openpyxl.styles.Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side
        ) 
        
        if delete_worksheet and sheetname in list(writer.sheets):
            del wb[sheetname]
            del writer.sheets[sheetname]
            
        if sheetname not in list(writer.sheets):
            df.to_excel(writer, sheetname, index=False, startrow=startrow, startcol=startcol, header=header)
            
        worksheet = writer.sheets[sheetname]
        
        if title is not None:
            worksheet.merge_cells(
                f"{get_int_excelcol(startcol+1)}{startrow}:"
                f"{get_int_excelcol(startcol+len(df.columns))}{startrow}"
            )
            
            cell = worksheet.cell(row=startrow, column=startcol+1)
            cell.value = title

            alignmeng_obj = copy(cell.alignment)
            alignmeng_obj.horizontal = 'center'
            alignmeng_obj.vertical = 'center'
            cell.alignment = alignmeng_obj
            cell.number_format = 'General'
            cell.border = border
            cell.fill = header_fill
            cell.font = header_font
            
        if header:
            for j, col in enumerate(df.columns):
                cell = worksheet.cell(row=startrow+1, column=startcol+j+1)
                cell.value = col
                
                alignmeng_obj = copy(cell.alignment)
                alignmeng_obj.horizontal = 'center'
                alignmeng_obj.vertical = 'center'
                cell.alignment = alignmeng_obj
                cell.number_format = 'General'
                cell.border = border
                cell.fill = header_fill
                cell.font = header_font
                
            startrow += 1
            
        # set cell values
        for i, (idx, row) in enumerate(df.iterrows()):
            for j, val in enumerate(row):
                if skip_nan and isinstance(val, float) and np.isnan(val):
                    continue
                
                cell = worksheet.cell(row=startrow+i+1, column=startcol+j+1)
                cell.value = val
                
                alignmeng_obj = copy(cell.alignment)
                alignmeng_obj.horizontal = 'center'
                alignmeng_obj.vertical = 'center'
                cell.alignment = alignmeng_obj
                cell.number_format = 'General'
                cell.border = border
                
                if j == 0 and index:
                    cell.number_format = 'General'
                    if color_index:
                        cell.fill = header_fill
                        cell.font = header_font
                elif isinstance(val, datetime):
                    cell.number_format = 'YYYY-MM-DD'
                elif isinstance(val, (int, float)):
                    col = df.columns.values[j]
                    if integer_columns is not None and col in integer_columns:
                        cell.number_format = '#,##0'
                    elif currency_columns is not None and col in currency_columns:
                        cell.number_format = '$#,##0'
                    elif pct_columns is not None and col in pct_columns:
                        cell.number_format = '#,##0.00%'
                    else:
                        cell.number_format = "0.00"
        
        # adjust columns widths
        for j, col in enumerate(df):
            series = df[col]
            max_len = min(
                64,
                max(
                    len(str(series.name)) + 3,
                    series.astype(str).map(len).max() + 3,
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(startcol+j+1)].width
                )
            )
            
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(startcol+j+1)].width = max_len
            
        worksheet.sheet_view.showGridLines = False
    
    # close file
    wb.close()
    
    
def get_int_excelcol(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n-1, 26)
        string = chr(65 + remainder) + string
        
    return string


def get_sheetnames(filename:str) -> List[str]:
    wb = openpyxl.load_workbook(filename)
    return wb.sheetnames


def get_nrows(filename:str, sheetname:str) -> int:
    wb = openpyxl.load_workbook(filename)
    
    worksheets = [ws for ws in wb.worksheets if ws.title == sheetname]
    if len(worksheets) == 0:
        raise FileNotFoundError(f"The sheentame '{sheetname}' does not exist in file '{filename}'")
    elif len(worksheets) == 1:
        return worksheets[0].max_row
    else:
        raise ValueError(f"The sheename '{sheetname}' is ambiguous.")
